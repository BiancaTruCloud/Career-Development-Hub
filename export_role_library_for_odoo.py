#!/usr/bin/env python3
"""
Career Development Hub — Role Library XLSX -> Odoo-ready export (pure Python)

Reads a Role Library workbook and writes 3 CSVs:
  1) skills.csv
  2) role_profiles.csv
  3) role_profile_lines.csv
"""

from __future__ import annotations

import argparse
import csv
import os
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

DEFAULT_INPUT_XLSX = "/mnt/data/SA_Role_Library_Municipality_and_Industries_with_NQF_SASCO (1).xlsx"
DEFAULT_SHEET_NAME = "Role Library"
DEFAULT_OUT_DIR = "out"

HARD_SKILL_COLS = [(f"Hard Skill {i}", f"Hard Skill {i} Level") for i in range(1, 7)]
SOFT_SKILL_COLS = [(f"Soft Skill {i}", f"Soft Skill {i} Level") for i in range(1, 7)]

LEVEL_MAP = {
    "beginner": 1,
    "intermediate": 2,
    "advanced": 3,
    "expert": 4,
}


def normalize_skill_name(name: str) -> str:
    s = (name or "").strip()
    return re.sub(r"\s+", " ", s)


def normalize_key(s: str) -> str:
    return normalize_skill_name(s).lower()


def norm_level(level_raw: Optional[str]) -> Optional[int]:
    if not level_raw:
        return None
    return LEVEL_MAP.get(str(level_raw).strip().lower())


def safe_str(v) -> str:
    return "" if v is None else str(v).strip()


@dataclass
class SkillRow:
    external_skill_key: str
    name: str
    skill_type: str
    category: str


@dataclass
class RoleProfileRow:
    external_role_id: str
    name: str
    role_title: str
    career_level: str
    sector: str
    industry: str
    department: str
    sub_department: str
    job_family: str
    role_description: str
    key_responsibilities: str
    psod_occupational_category: str
    psod_skill_level: str
    nqf_band: str
    recommended_nqf_levels: str
    sasko_major_group: str
    sasko_skill_level: str
    sasko_unit_group_code: str
    import_source: str
    last_imported_on: str


@dataclass
class RoleProfileLineRow:
    external_role_id: str
    external_skill_key: str
    skill_name: str
    skill_type: str
    target_level_seq: int
    is_required: bool


def read_sheet_as_dicts(wb, sheet_name: str) -> List[Dict[str, str]]:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [safe_str(h) for h in rows[0]]
    header_index = {h: i for i, h in enumerate(headers) if h}
    data: List[Dict[str, str]] = []
    for r in rows[1:]:
        row_dict: Dict[str, str] = {}
        for h, idx in header_index.items():
            row_dict[h] = "" if idx >= len(r) or r[idx] is None else str(r[idx]).strip()
        if any(v for v in row_dict.values()):
            data.append(row_dict)
    return data


def detect_column_name(possible: List[str], row: Dict[str, str]) -> Optional[str]:
    keys = set(row.keys())
    for c in possible:
        if c in keys:
            return c
    return None


def export_role_library(input_xlsx: str, sheet_name: str, out_dir: str) -> Tuple[int, int, int]:
    if not os.path.exists(input_xlsx):
        raise FileNotFoundError(f"Input XLSX not found: {input_xlsx}")

    wb = load_workbook(input_xlsx, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    records = read_sheet_as_dicts(wb, sheet_name)
    if not records:
        raise ValueError(f"No data rows found in sheet '{sheet_name}'")

    role_id_col = detect_column_name(["Role ID", "RoleID", "Role Id"], records[0])
    role_title_col = detect_column_name(["Role Title", "Title", "Role"], records[0])
    career_level_col = detect_column_name(["Career Level", "Level"], records[0])

    sector_col = detect_column_name(["Sector"], records[0])
    industry_col = detect_column_name(["Industry"], records[0])
    dept_col = detect_column_name(["Department"], records[0])
    sub_dept_col = detect_column_name(["Sub-Department", "Sub Department"], records[0])
    job_family_col = detect_column_name(["Job Family"], records[0])
    role_desc_col = detect_column_name(["Role Description", "Description"], records[0])
    resp_col = detect_column_name(["Key Responsibilities", "Responsibilities"], records[0])

    psod_occ_col = detect_column_name(["PSOD Occupational Category"], records[0])
    psod_skill_col = detect_column_name(["PSOD Skill Level"], records[0])
    nqf_band_col = detect_column_name(["NQF Band"], records[0])
    nqf_rec_col = detect_column_name(["Recommended NQF Level(s)", "Recommended NQF Levels"], records[0])
    sasko_major_col = detect_column_name(["SASCO Major Group"], records[0])
    sasko_skill_col = detect_column_name(["SASCO Skill Level"], records[0])
    sasko_unit_col = detect_column_name(["SASCO Unit Group Code"], records[0])

    if not role_title_col:
        raise ValueError("Could not find required column 'Role Title' (or fallback).")

    os.makedirs(out_dir, exist_ok=True)

    skills: Dict[str, SkillRow] = {}
    role_profiles: Dict[str, RoleProfileRow] = {}
    role_lines: List[RoleProfileLineRow] = []

    import_source = os.path.basename(input_xlsx)
    imported_on = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")

    for row in records:
        role_title = safe_str(row.get(role_title_col, ""))
        if not role_title:
            continue

        career_level = safe_str(row.get(career_level_col, "")) if career_level_col else ""
        external_role_id = safe_str(row.get(role_id_col, "")) if role_id_col else ""
        if not external_role_id:
            external_role_id = normalize_key(f"{role_title}::{career_level}")[:180] or normalize_key(role_title)[:180]

        if external_role_id not in role_profiles:
            name = role_title if not career_level else f"{role_title} ({career_level})"
            role_profiles[external_role_id] = RoleProfileRow(
                external_role_id=external_role_id,
                name=name,
                role_title=role_title,
                career_level=career_level,
                sector=safe_str(row.get(sector_col, "")) if sector_col else "",
                industry=safe_str(row.get(industry_col, "")) if industry_col else "",
                department=safe_str(row.get(dept_col, "")) if dept_col else "",
                sub_department=safe_str(row.get(sub_dept_col, "")) if sub_dept_col else "",
                job_family=safe_str(row.get(job_family_col, "")) if job_family_col else "",
                role_description=safe_str(row.get(role_desc_col, "")) if role_desc_col else "",
                key_responsibilities=safe_str(row.get(resp_col, "")) if resp_col else "",
                psod_occupational_category=safe_str(row.get(psod_occ_col, "")) if psod_occ_col else "",
                psod_skill_level=safe_str(row.get(psod_skill_col, "")) if psod_skill_col else "",
                nqf_band=safe_str(row.get(nqf_band_col, "")) if nqf_band_col else "",
                recommended_nqf_levels=safe_str(row.get(nqf_rec_col, "")) if nqf_rec_col else "",
                sasko_major_group=safe_str(row.get(sasko_major_col, "")) if sasko_major_col else "",
                sasko_skill_level=safe_str(row.get(sasko_skill_col, "")) if sasko_skill_col else "",
                sasko_unit_group_code=safe_str(row.get(sasko_unit_col, "")) if sasko_unit_col else "",
                import_source=import_source,
                last_imported_on=imported_on,
            )

        def process_skill_pairs(pairs: List[Tuple[str, str]], skill_type: str) -> None:
            for skill_col, level_col in pairs:
                skill_name_raw = safe_str(row.get(skill_col, ""))
                if not skill_name_raw:
                    continue
                skill_name = normalize_skill_name(skill_name_raw)
                if not skill_name:
                    continue

                level_seq = norm_level(row.get(level_col, ""))
                if level_seq is None:
                    continue

                external_skill_key = normalize_key(skill_name)
                if external_skill_key not in skills:
                    skills[external_skill_key] = SkillRow(
                        external_skill_key=external_skill_key,
                        name=skill_name,
                        skill_type=skill_type,
                        category="Uncategorized",
                    )

                role_lines.append(
                    RoleProfileLineRow(
                        external_role_id=external_role_id,
                        external_skill_key=external_skill_key,
                        skill_name=skill_name,
                        skill_type=skill_type,
                        target_level_seq=level_seq,
                        is_required=True,
                    )
                )

        process_skill_pairs(HARD_SKILL_COLS, "hard")
        process_skill_pairs(SOFT_SKILL_COLS, "soft")

    dedup: Dict[Tuple[str, str], RoleProfileLineRow] = {}
    for line in role_lines:
        key = (line.external_role_id, line.external_skill_key)
        if key not in dedup or line.target_level_seq > dedup[key].target_level_seq:
            dedup[key] = line
    role_lines = list(dedup.values())

    skills_path = os.path.join(out_dir, "skills.csv")
    roles_path = os.path.join(out_dir, "role_profiles.csv")
    lines_path = os.path.join(out_dir, "role_profile_lines.csv")

    with open(skills_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["external_skill_key", "name", "skill_type", "category"])
        for s in sorted(skills.values(), key=lambda x: (x.skill_type, x.name.lower())):
            w.writerow([s.external_skill_key, s.name, s.skill_type, s.category])

    with open(roles_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "external_role_id", "name", "role_title", "career_level",
            "sector", "industry", "department", "sub_department", "job_family",
            "role_description", "key_responsibilities",
            "psod_occupational_category", "psod_skill_level",
            "nqf_band", "recommended_nqf_levels",
            "sasko_major_group", "sasko_skill_level", "sasko_unit_group_code",
            "import_source", "last_imported_on",
        ])
        for r in sorted(role_profiles.values(), key=lambda x: x.name.lower()):
            w.writerow([
                r.external_role_id, r.name, r.role_title, r.career_level,
                r.sector, r.industry, r.department, r.sub_department, r.job_family,
                r.role_description, r.key_responsibilities,
                r.psod_occupational_category, r.psod_skill_level,
                r.nqf_band, r.recommended_nqf_levels,
                r.sasko_major_group, r.sasko_skill_level, r.sasko_unit_group_code,
                r.import_source, r.last_imported_on,
            ])

    with open(lines_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "external_role_id", "external_skill_key", "skill_name", "skill_type", "target_level_seq", "is_required",
        ])
        for l in sorted(role_lines, key=lambda x: (x.external_role_id, x.skill_type, x.skill_name.lower())):
            w.writerow([
                l.external_role_id,
                l.external_skill_key,
                l.skill_name,
                l.skill_type,
                l.target_level_seq,
                "1" if l.is_required else "0",
            ])

    return len(skills), len(role_profiles), len(role_lines)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Export Role Library XLSX to Odoo-friendly CSV files.")
    parser.add_argument("--input", default=DEFAULT_INPUT_XLSX, help="Path to source XLSX file")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Sheet name in workbook")
    parser.add_argument("--out", default=DEFAULT_OUT_DIR, help="Output directory for CSV files")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    skills_count, roles_count, lines_count = export_role_library(args.input, args.sheet, args.out)
    print("Done.")
    print(f"- Skills:        {args.out}/skills.csv ({skills_count})")
    print(f"- Role Profiles: {args.out}/role_profiles.csv ({roles_count})")
    print(f"- Profile Lines: {args.out}/role_profile_lines.csv ({lines_count})")


if __name__ == "__main__":
    main()
